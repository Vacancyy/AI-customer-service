"""
生成客服评审表：从知识库抽样 → RAG生成回答 → 导出Excel让客服评分
"""

import os
import sys

PROJECT_ROOT = os.environ.get('PROJECT_ROOT', os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'scripts'))

import json
import random
import time
from ai_qa_system_v2 import ImprovedQASystem

QA_PATH = os.path.join(PROJECT_ROOT, '05_analyze/reports/知识库_优化版.json')
OUTPUT_PATH = os.path.join(PROJECT_ROOT, '05_analyze/reports/客服评审表.xlsx')

def main():
    # 加载知识库用于选题
    with open(QA_PATH, 'r') as f:
        data = json.load(f)

    # 分组抽样
    random.seed(42)
    groups = {}
    for i, qa in enumerate(data):
        cat = qa.get('primary_category', '其他')
        if cat not in groups:
            groups[cat] = []
        groups[cat].append((i, qa))

    plan = {
        '产品信息': 6,
        '理赔流程': 6,
        '保障范围': 6,
        '条款解释': 4,
        '其他问题': 3,
        '理赔材料': 3,
        '退保流程': 2,
    }

    selected = []
    for cat, count in plan.items():
        items = groups.get(cat, [])
        random.shuffle(items)
        for idx, qa in items[:count]:
            selected.append((idx, qa))
    selected.sort(key=lambda x: x[0])

    # 初始化RAG系统
    print("初始化RAG系统...")
    system = ImprovedQASystem()

    # 逐题生成回答
    results = []
    for i, (idx, qa) in enumerate(selected):
        question = qa['std_question']
        print(f"[{i+1}/{len(selected)}] {question[:40]}...", end=' ', flush=True)

        start = time.time()
        answer = system.get_answer(question)
        elapsed = time.time() - start

        # 检索到的知识库原文
        matched = system.search_knowledge(question, top_k=1)
        kb_original = matched[0]['qa']['answer'] if matched else ''
        match_score = matched[0]['score'] if matched else 0

        results.append({
            '序号': i + 1,
            '分类': qa.get('primary_category', ''),
            '问题': question,
            '知识库标准答案': qa['answer'],
            'AI回答': answer,
            '检索命中条目': matched[0]['qa']['std_question'] if matched else '',
            '检索相似度': f'{match_score:.3f}',
            '响应时间(秒)': f'{elapsed:.1f}',
            '回答是否正确': '',
            '有无遗漏': '',
            '有无编造': '',
            '客服备注': '',
        })
        print(f"✓ ({elapsed:.1f}s)")

    # 导出Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客服评审"

    # 表头
    headers = list(results[0].keys())
    header_fills = {
        '回答是否正确': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        '有无遗漏': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        '有无编造': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        '客服备注': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    }

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        if header in header_fills:
            cell.fill = header_fills[header]

    # 写数据
    for row_idx, item in enumerate(results, 2):
        for col, header in enumerate(headers, 1):
            value = item[header]
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
            if header in header_fills:
                cell.fill = header_fills[header]

    # 列宽
    col_widths = {
        'A': 5,   # 序号
        'B': 10,  # 分类
        'C': 30,  # 问题
        'D': 50,  # 知识库标准答案
        'E': 50,  # AI回答
        'F': 30,  # 检索命中条目
        'G': 10,  # 检索相似度
        'H': 10,  # 响应时间
        'I': 12,  # 回答是否正确
        'J': 12,  # 有无遗漏
        'K': 12,  # 有无编造
        'L': 20,  # 客服备注
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加说明sheet
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["客服评审表填写说明"],
        [""],
        ["1. 评审目的", "验证AI客服回答是否准确，是否可以直接上线使用"],
        ["2. 填写方式", "在黄色列中填写评审意见"],
        ["3. 回答是否正确", "填写：✓ 正确 / ✗ 错误 / △ 部分正确"],
        ["4. 有无遗漏", "填写：无遗漏 / 有遗漏（注明遗漏内容）"],
        ["5. 有无编造", "填写：无编造 / 有编造（注明编造内容）"],
        ["6. 客服备注", "补充说明，如建议修改的回答等"],
        ["7. 知识库标准答案", "供对照参考，AI应基于此内容回答"],
        ["8. 检索相似度", "越高说明检索越精准，>0.8为良好"],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 50
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

    wb.save(OUTPUT_PATH)
    print(f"\n评审表已生成: {OUTPUT_PATH}")
    print(f"共 {len(results)} 道题，请交给客服评审")


if __name__ == '__main__':
    main()
